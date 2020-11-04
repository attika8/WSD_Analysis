import dictionary as dictionary
import requests
import urllib3
from pprint import pprint
import xml.etree.ElementTree as ET
from lxml import etree
import re
import pandas as pd
import xmltodict
import json
from openpyxl import load_workbook
import time
from tqdm import tqdm
import pandas as pd
import getpass
import dpath.util
import functools
from collections import abc
from collections.abc import MutableMapping
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
'''
import smtplib, ssl

port = 587  # For starttls
smtp_server = "smtp.gmail.com"
sender_email = "attila.labai@gmail.com"
receiver_email = "attika8@gmail.com"
mypass = getpass.getpass()
print(mypass)

password = getpass.getpass("PASSWORD : ")
#password = getpass.getpass('Type your password and press enter:')
message = """\
Subject: Hi there

This message is sent from Python."""

context = ssl.create_default_context()
with smtplib.SMTP(smtp_server, port) as server:
    server.ehlo()  # Can be omitted
    server.starttls(context=context)
    server.ehlo()  # Can be omitted
    server.login(sender_email, password)
    server.sendmail(sender_email, receiver_email, message)
'''

utf8_parser = etree.XMLParser(encoding='utf-8')

start = time.time()
#601 master
#322 omnibus
#341 tcs

url = "https://citiuat.wallstreetdocs.local/api/rest/termSheet/list/structure/601.xml"
response = requests.get(url,auth=('admin','citi@w$d'),verify=False)

myTree = etree.fromstring(response.content, parser=utf8_parser)
numberOfAllTermsheets = len(myTree[0].getchildren())
print('All Termsheet in UAT: ' + str(numberOfAllTermsheets))

def getListOfTermsheetIDs(RegressionTesting=True):
    listOfIDs = []

    for termsheet in myTree.iter():
        if(RegressionTesting):
            try:
                if re.match('REGS.*', termsheet.attrib['name']) \
                    or re.match('.*DPWAPPROVED.*',termsheet.attrib['name']) \
                    or re.match('.*SECTS.*',termsheet.attrib['name']) \
                    or re.match('REG TEST SWAP TS', termsheet.attrib['name']) \
                    or re.match('.*MLDMS.*',termsheet.attrib['name']) \
                    or re.match('APACDig.*',termsheet.attrib['name']) \
                    or re.match('JPNMUMS.*',termsheet.attrib['name']) \
                    or re.match('.*APACSnow.*',termsheet.attrib['name']) \
                    or re.match('.*APACTERMSHEET.*',termsheet.attrib['name']) \
                    or re.match('.*APACFCA.*',termsheet.attrib['name']) \
                    or re.match('.*APACDRA.*',termsheet.attrib['name']) \
                    or re.match('.*JPNMizuho.*',termsheet.attrib['name']) \
                    or re.match('Cert.*', termsheet.attrib['name']) \
                    or re.match('Infosheet.*', termsheet.attrib['name']) \
                    or re.match('.*TCS Reg.*',termsheet.attrib['name']) \
                    or re.match('.*TCSTS*.',termsheet.attrib['name']) \
                    or re.match('.*TCSPS.*',termsheet.attrib['name']):
                        listOfIDs.append(termsheet.attrib['id'])
                        #print(termsheet.attrib['name'])
            except KeyError:
                pass
        else:
            try:
                listOfIDs.append(termsheet.attrib['id'])
            except KeyError:
                pass

    #listOfIDs.append(7987)
    #listOfIDs.append(39823)
    #listOfIDs.append(32012)
    #listOfIDs.append(32047)

    #if termsheet.attrib['name'] == '20HEQ002882':
    #    print('this')

    #listOfIDs = []
    #for i in range(48382,48410):
        #if '20HEQ002882' == termsheet.attrib['name']:
        #    print('20HEQ002'+str(i) + ": " + str(termsheet.attrib['id']))
    #        listOfIDs.append(i)

    print('Number of IDs found: ' + str(len(listOfIDs)))
    return listOfIDs


def handleTSData(id):

    dictionary = []
    url = "https://citiuat.wallstreetdocs.local/api/rest/termSheet/" + str(id)
    XML = requests.get(url, auth=('admin', 'citi@w$d'), verify=False)
    #postURL = "https://citiuat.wallstreetdocs.local/api/rest/termSheet/"+str(id)
    myTree = etree.fromstring(XML.content, parser=utf8_parser)
    saveResponse = requests.post(url, data=XML.content, auth=('admin', 'citi@w$d'), verify=False)
    #print(saveResponse)
    print('resaved ' + str(id))
    dictionary.append(('id',str(myTree[0].text)))
    print('Capturing XML id: ' + str(myTree[0].text))
    dictionary.append(('TS_Name',str(myTree[1].text)))
    for elem in myTree.iter():
        if elem.tag == 'item':
            try:
                itemName = elem.attrib['name']
                itemTag = elem[0].tag
                key = itemName
            except KeyError:
                pass

            if elem[0].tag == "compound":
                for numberTags in elem[0]:
                    if numberTags[0].attrib['name'] == 'Range?':
                        numberName = elem.attrib['name']
                        for values in numberTags:
                            if values.attrib['name'] == 'Range?':
                                rangeID = values.attrib['id']

                            if int(rangeID) == 0:
                                if values.attrib['name'] == 'Value':
                                    value = values[0].text

                                dictionary.append((numberName, value))

                            else:
                                if values.attrib['name'] == 'Lower end':
                                    lowerEndValue = values[0].text
                                if values.attrib['name'] == 'Upper end':
                                    upperEndValue = values[0].text

                                dictionary.append((numberName, (lowerEndValue,upperEndValue)))
                    else:
                        continue
            else:
                itemTag = elem[0].tag
                #print(itemTag)
                for node in elem:
                    #print(itemTag)
                    if node.tag == "array":
                        keyCount = 0
                        arrayValues = []

                        for arrayNode in node.findall('items/item/compound/items/item'):
                            arrayKey = arrayNode.attrib['name']
                            if arrayNode.tag == "item":
                                if arrayNode.tag == "master":
                                    continue
                                else:
                                    try:
                                        arrayValues = []
                                        indexValue = int(arrayNode[0].attrib['index'])
                                        #print(indexValue)
                                        try:
                                            arrayText = arrayNode[0][0][indexValue].text
                                            node.tag = "array"
                                            arrayValues.append(arrayText)
                                        except KeyError:
                                            pass
                                    except KeyError:
                                        pass
                                #print(arrayNode[1][indexValue].text)
                        #print(node.tag)
                        #print(arrayValues)
                        #if arrayKey == itemName:
                        #    keyCount = keyCount+1

                        #if keyCount == 0:
                        dictionary.append((key,','.join(map(str, arrayValues))))
                        #dictionary.append((key, arrayValues))
                        #print('this is an array - list')

                    elif node.tag == "select":
                        if int(str(node.getroottree().getpath(node)).find('array')) > 0:
                            break
                        else:
                            indexValue = int(node.attrib['index'])
                            try:
                                name = node.attrib['name']
                            except KeyError:
                                pass
                            value = node[0][indexValue].text
                            dictionary.append((key,value)) #this is where array select is defined extra
                    #print(node.tag, node.attrib)
                    elif node.tag == "text" or node.tag == "date":
                        tagValue = node.text
                        dictionary.append((key, tagValue))
                        #print('this is a text field')
                    elif node.tag == "dictionary":
                        indexValue = int(node.attrib['index'])
                        value = node[0][indexValue].text
                        dictionary.append((key, value))
    return dictionary


def handleData():
    fullList = []
    for id in tqdm(getListOfTermsheetIDs()):
        singleTSData = handleTSData(id)
        fullList.append(singleTSData)

    #pprint(singleTSData)

    fullDict = {item[0][1]:dict(item) for item in fullList}
    return fullDict

def resaveTermsheets():
    for id in tqdm(getListOfTermsheetIDs(RegressionTesting=False)):
        url = "https://citiuat.wallstreetdocs.local/api/rest/termSheet/" + str(id)
        XML = requests.get(url, auth=('admin', 'citi@w$d'), verify=False)
        requests.post(url, data=XML.content, auth=('admin', 'citi@w$d'), verify=False)
        #print('resaved ' + str(id))


def regexlistcols(dataDf, string):
    ''' Utility Took
    :param dataDf: DataFrame with many columns
    :param string: string to be included somewhere in the column name
    :return: prints column names in dataDf matching regex
    '''
    for s in dataDf.columns.to_list():
        if re.match('\w*' + string + '\w*', s):
            print(s)

def outputToExcel(outputDF, outputFile, outputSheet, resetFile=True):
    '''
    Outputs outputDF to outputSheet in outputFile
    :param resetFile: when True removes existing sheets in outputFile
    :param outputDF: DataFrame to output
    :param outputFile: File to output to
    :param outputSheet: Sheet to output to
    :return: nothing
    '''

    print('Started outputting to Excel')

    book = load_workbook(outputFile)

    if resetFile:
        # Remove existing sheets
        for sheet in book.worksheets:
            book.remove(sheet)
        book.create_sheet('PlaceholderSheetToEnableRemove')
        book.save(outputFile)
        book = load_workbook(outputFile)

    writer = pd.ExcelWriter(outputFile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    outputDF.to_excel(writer, outputSheet)

    if resetFile:
        book.remove(book['PlaceholderSheetToEnableRemove'])
    print('Done outputting to Excel')
    writer.save()


#output = handleData()
#tsInfo = pd.DataFrame.from_dict(dict(output)).T
#print(tsInfo)

resaveTermsheets()

#print(regexlistcols(tsInfo,'roduc'))

#GET the list of taglibs
#https://citi.wallstreetdocs.local/api/rest/taglib/list/structure/601

'''''''''''
"codeBaseConfig": [
			{
				"title": "Taglibs",
				"fsScheme": "taglibFs",
				"fileExtension": "php",
				"search": {
					"method": "GET",
					"path": "/api/rest/taglib/list/structure/{{id}}.json"
				},
				"revisionSearch": {
					"method": "GET",
					"path": "/api/rest/taglib/get/{{id}}/versions.json"
				},
				"getRevision": {
					"method": "GET",
					"path": "/api/rest/taglib/get/{{id}}?revision={{revisionId}}"
				},
				"get": {
					"method": "GET",
					"path": "/api/rest/taglib/get/{{id}}"
				},
				"update": {
					"method": "POST",
					"path": "/api/rest/taglib/put/{{id}}"
				},
				"create": {
					"method": "POST",
					"path": "/api/rest/taglib/create/{{parentId}}?tagName={{name}}"
				},
				"delete": {
					"method": "POST",
					"path": "/api/rest/taglib/delete/{{id}}"
				}
			},
			{
				"title": "Codelibs",
				"fsScheme": "codelibFs",
				"fileExtension": "php",
				"search": {
					"method": "GET",
					"path": "/api/rest/codeLibrary/list/structure/{{id}}.json"
				},
				"revisionSearch": {
					"method": "GET",
					"path": "/api/rest/codeLibrary/get/{{id}}/versions.json"
				},
				"getRevision": {
					"method": "GET",
					"path": "/api/rest/codeLibrary/get/{{id}}?revision={{revisionId}}"
				},
				"get": {
					"method": "GET",
					"path": "/api/rest/codeLibrary/get/{{id}}"
				},
				"update": {
					"method": "POST",
					"path": "/api/rest/codeLibrary/put/{{id}}"
				},
				"create": {
					"method": "POST",
					"path": "/api/rest/codeLibrary/create/{{parentId}}?codeLibName={{name}}"
				},
				"delete": {
					"method": "POST",
					"path": "/api/rest/codeLibrary/delete/{{id}}"
				}
			}
		]
'''''''''
end = time.time()
print(float(end - start))
